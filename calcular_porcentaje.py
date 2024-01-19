import pandas as pd
from openpyxl import load_workbook
def calcularPorcent(data,porcent,final_path,condition,cal):

    df = pd.DataFrame(data)
    valores_unicos_lista = df['year'].unique().tolist()
    meses_unicos = []
    for year in valores_unicos_lista:
        c1 = df['year'] == year
        datos_seleccionados = df[c1]
        meses_unicos.append({'meses':datos_seleccionados['mes'].unique().tolist(),'year':year})
    longitud_meses = 0
    for lon in meses_unicos:
        longitud_meses += len(lon['meses'])
    size = df.shape[0]
    total_porcent = 24
    muestra_generada = []
    if condition == 1:
        total_porcent = round(int(size)*1)
        muestra_generada = data
    elif condition == 2:
        N_poblation = size
        error = 0.05 
        confianza = 1.96 
        total_porcent = (N_poblation *(confianza**2)*0.5*0.5)/((error**2)*(N_poblation-1)+(confianza**2)*0.5*0.5)
      
        for dic in meses_unicos:
            c2 = df['year'] == dic['year']
            data_frame_year = df[c2]
            for val in dic['meses']:
                c3 = data_frame_year['mes'] == val
                df_mes = data_frame_year[c3]
                elementos_aleatorios = df_mes.sample(n=ctd_mes, replace=False)
                aux = elementos_aleatorios.to_dict(orient='records')
                muestra_generada.extend(aux)
    elif condition == 3:
        total_porcent = round(int(size)*porcent)
        ctd_mes = round(total_porcent/longitud_meses)
        for dic in meses_unicos:
            c2 = df['year'] == dic['year']
            data_frame_year = df[c2]
            for val in dic['meses']:
                c3 = data_frame_year['mes'] == val
                df_mes = data_frame_year[c3]
                elementos_aleatorios = df_mes.sample(n=ctd_mes, replace=False)
                aux = elementos_aleatorios.to_dict(orient='records')
                muestra_generada.extend(aux)
    

    if cal == 20:
        workbook = load_workbook('./plantillas/cal20_auditar.xlsx')
        sheet = workbook.active
        aux_init = 3
        
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['year']
            sheet[f'B{aux_init}'] = i['mes']
            sheet[f'C{aux_init}'] = i['dia']
            sheet[f'D{aux_init}'] = i['subestacion']
            sheet[f'E{aux_init}'] = i['geo_x']
            sheet[f'F{aux_init}'] = i['geo_y']
            sheet[f'G{aux_init}'] = i['provincia']
            sheet[f'H{aux_init}'] = i['canton']
            sheet[f'K{aux_init}'] = i['registros']
            sheet[f'L{aux_init}'] = i['fase_av']
            sheet[f'M{aux_init}'] = i['fase_bv']
            sheet[f'N{aux_init}'] = i['fase_cv']
            sheet[f'T{aux_init}'] = i['observaciones']
            aux_init += 1
        
        workbook.save(final_path)
        workbook.close()
    elif cal == 30:
        workbook = load_workbook('./plantillas/cal30_auditar.xlsx')
        sheet = workbook.active
        aux_init = 4
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['year']
            sheet[f'B{aux_init}'] = i['mes']
            sheet[f'C{aux_init}'] = i['dia']
            sheet[f'D{aux_init}'] = i['codigo']
            sheet[f'E{aux_init}'] = i['tipo']
            sheet[f'F{aux_init}'] = i['subestacion']
            sheet[f'G{aux_init}'] = i['alimentador']
            sheet[f'H{aux_init}'] = i['fases']
            sheet[f'I{aux_init}'] = i['ff']
            sheet[f'J{aux_init}'] = i['fn']
            sheet[f'L{aux_init}'] = i['registros']
            sheet[f'M{aux_init}'] = i['fase_av']
            sheet[f'N{aux_init}'] = i['fase_apst']
            sheet[f'O{aux_init}'] = i['fase_avthd']
            sheet[f'P{aux_init}'] = i['fase_cv']
            sheet[f'Q{aux_init}'] = i['fase_cpst']
            sheet[f'R{aux_init}'] = i['fase_cvthd']
            sheet[f'S{aux_init}'] = i['fase_bv']
            sheet[f'T{aux_init}'] = i['fase_bpst']
            sheet[f'U{aux_init}'] = i['fase_bvthd']
            sheet[f'V{aux_init}'] = i['desequilibrio']
            sheet[f'AI{aux_init}'] = i['observaciones']
            aux_init += 1
        workbook.save(final_path)
        workbook.close()
    elif cal == 40:
        workbook = load_workbook('./plantillas/cal40_auditar.xlsx')
        sheet = workbook.active
        aux_init = 3
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['year']
            sheet[f'B{aux_init}'] = i['mes']
            sheet[f'C{aux_init}'] = i['dia']
            sheet[f'D{aux_init}'] = i['codigo']
            sheet[f'E{aux_init}'] = i['tipo']
            sheet[f'F{aux_init}'] = i['provincia']
            sheet[f'G{aux_init}'] = i['canton']
            sheet[f'H{aux_init}'] = i['subestacion']
            sheet[f'I{aux_init}'] = i['alimentador']
            sheet[f'J{aux_init}'] = i['transformador']
            sheet[f'K{aux_init}'] = i['fases']
            sheet[f'L{aux_init}'] = i['ff']
            sheet[f'M{aux_init}'] = i['fn']
            sheet[f'O{aux_init}'] = i['registros']
            sheet[f'P{aux_init}'] = i['fase_av']
            sheet[f'Q{aux_init}'] = i['fase_bv']
            sheet[f'R{aux_init}'] = i['fase_cv']
            sheet[f'W{aux_init}'] = i['observaciones']
        
            aux_init += 1
        workbook.save(final_path)
        workbook.close()
    elif cal == 50:
        workbook = load_workbook('./plantillas/cal50_auditar.xlsx')
        sheet = workbook.active
        aux_init = 4
        for i in muestra_generada:
            sheet[f'A{aux_init}'] = i['year']
            sheet[f'B{aux_init}'] = i['mes']
            sheet[f'C{aux_init}'] = i['dia']
            sheet[f'D{aux_init}'] = i['codigo']
            sheet[f'E{aux_init}'] = i['provincia']
            sheet[f'F{aux_init}'] = i['canton']
            sheet[f'G{aux_init}'] = i['subestacion']
            sheet[f'H{aux_init}'] = i['alimentador']
            sheet[f'I{aux_init}'] = i['ff']
            sheet[f'J{aux_init}'] = i['fn']
            sheet[f'L{aux_init}'] = i['registros']
            sheet[f'M{aux_init}'] = i['fase_av']
            sheet[f'N{aux_init}'] = i['fase_apst']
            sheet[f'O{aux_init}'] = i['fase_avthd']
            sheet[f'P{aux_init}'] = i['fase_bv']
            sheet[f'Q{aux_init}'] = i['fase_bpst']
            sheet[f'R{aux_init}'] = i['fase_bvthd']
            sheet[f'S{aux_init}'] = i['fase_cv']
            sheet[f'T{aux_init}'] = i['fase_cpst']
            sheet[f'U{aux_init}'] = i['fase_cvthd']
            sheet[f'V{aux_init}'] = i['desequilibrio']
            sheet[f'AI{aux_init}'] = i['observaciones']

            aux_init += 1
        workbook.save(final_path)
        workbook.close()




